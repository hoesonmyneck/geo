"""Выгружает всю pop_dwelling в Excel (.xlsx) для визуальной проверки.

Строк > 1 048 576 (лимит листа Excel) → бьём на несколько листов, чтобы влезли ВСЕ.
Потоково: server-side курсор + write_only openpyxl (память не пухнет).

Запуск (ХОСТ): python backend/worker/export_pop_dwelling_xlsx.py [out.xlsx]
"""
from __future__ import annotations
import sys
import psycopg
from openpyxl import Workbook

DSN = "host=localhost port=5432 dbname=geo user=geo password=geopassword123"
XLSX_MAX = 1_048_576           # лимит строк на лист (включая заголовок)
HEADER = ["dwelling_id", "kind", "geocode_addr", "людей (rca_count)", "lat", "lon"]


def main() -> None:
    out = sys.argv[1] if len(sys.argv) > 1 else "pop_dwelling.xlsx"
    wb = Workbook(write_only=True)
    ws = wb.create_sheet("Жилища_1")
    ws.append(HEADER)
    sheet_rows = 1              # заголовок
    sheet_no = 1
    total = 0

    conn = psycopg.connect(DSN)
    cur = conn.cursor(name="exp")           # server-side (стрим)
    cur.itersize = 20000
    cur.execute("""
        SELECT dwelling_id, kind, geocode_addr, rca_count, lat, lon
          FROM pop_dwelling ORDER BY dwelling_id
    """)
    for row in cur:
        if sheet_rows >= XLSX_MAX:           # лист полон — новый
            sheet_no += 1
            ws = wb.create_sheet(f"Жилища_{sheet_no}")
            ws.append(HEADER)
            sheet_rows = 1
        ws.append(list(row))
        sheet_rows += 1
        total += 1
        if total % 200000 == 0:
            print(f"  {total:,} строк...", flush=True)

    cur.close(); conn.close()
    print(f"сохраняю {total:,} строк на {sheet_no} лист(ах) -> {out}", flush=True)
    wb.save(out)
    print("готово")


if __name__ == "__main__":
    main()
