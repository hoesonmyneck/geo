"""
Импорт кандасов из нового xlsx (kandas_perso).

Структура xlsx:
- row 1: разделители секций
- row 2: русские заголовки
- row 3: английские коды колонок (используем их)
- row 4+: данные

Логика:
- APP_ID = идентификатор семьи (один на семью)
- APPLICANT = 1 → главный кандас, 0 → член семьи
- Один и тот же главный кандас (IIN) может быть в НЕСКОЛЬКИХ строках,
  каждая = новая запись об образовании/работе.
- Члены семьи (APPLICANT=0) идут в extra.family главного кандаса
  в текущем формате [{role, fio, iin}] — БЕЗ новых полей.

Запуск:
    docker compose cp kandas_perso\\ \\(2\\)\\ \\(1\\).xlsx backend:/app/data/input/kandas_perso.xlsx
    docker compose exec backend bash -c "cd /app && python worker/seed_kandas_xlsx.py /app/data/input/kandas_perso.xlsx"
"""
from __future__ import annotations

import asyncio
import sys
from collections import defaultdict
from pathlib import Path

sys.path.insert(0, "/app")

import openpyxl
from sqlalchemy import select

from app.db.session import AsyncSessionLocal
from app.db.models import Kandas


# Колонка с английским кодом → имя поля.
# Строка с кодами определяется автоматически (по наличию IIN+APP_ID),
# т.к. в разных выгрузках шапка занимает 2 или 3 строки.
def _read_xlsx(path: Path) -> list[dict]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    # Ищем строку английских кодов среди первых 5 строк
    hdr_row = None
    for r in range(1, 6):
        vals = {str(ws.cell(r, c).value).strip() for c in range(1, ws.max_column + 1) if ws.cell(r, c).value}
        if "IIN" in vals and "APP_ID" in vals:
            hdr_row = r
            break
    if hdr_row is None:
        raise ValueError("Не нашёл строку английских кодов (IIN/APP_ID) в первых 5 строках")
    # Читаем с col1: в разных выгрузках APP_ID бывает в 1-й колонке (без ведущего
    # пустого столбца) или во 2-й. Пустой ведущий столбец даст безвредный ключ.
    headers = {col: ws.cell(hdr_row, col).value for col in range(1, ws.max_column + 1) if ws.cell(hdr_row, col).value}
    rows: list[dict] = []
    for r in range(hdr_row + 1, ws.max_row + 1):
        d: dict = {}
        for col, name in headers.items():
            val = ws.cell(r, col).value
            if val is not None:
                # дату → ISO-строка
                if hasattr(val, "isoformat"):
                    val = val.isoformat() if "T" in str(val) or " " in str(val) else val.strftime("%d.%m.%Y")
                    # упрощённо: только дата если время 00:00:00
                    if hasattr(ws.cell(r, col).value, "strftime"):
                        v = ws.cell(r, col).value
                        if v.hour == 0 and v.minute == 0 and v.second == 0:
                            val = v.strftime("%d.%m.%Y")
                        else:
                            val = v.strftime("%d.%m.%Y %H:%M:%S")
                d[name] = val
        if not d.get("IIN"):
            continue
        # APP_ID нет → пропускаем
        if not d.get("APP_ID"):
            continue
        rows.append(d)
    return rows


def _s(v) -> str | None:
    """Безопасно в строку, None если пусто/прочерк."""
    if v is None:
        return None
    s = str(v).strip()
    if not s or s == "-":
        return None
    return s


def _iin(v) -> str | None:
    """Нормализация ИИН: всегда 12 цифр строкой (xlsx может хранить как число)."""
    if v is None:
        return None
    s = str(v).strip()
    if "." in s:
        # 1.15524E+12 → беда (теряем точность). По возможности парсим как int.
        try:
            s = str(int(float(s)))
        except ValueError:
            pass
    s = s.split(".")[0]
    # 11 цифр (число потеряло ведущий 0) — добиваем
    if s.isdigit() and len(s) == 11:
        s = "0" + s
    if s.isdigit() and len(s) == 12:
        return s
    return s if s else None


def _build_full_name(row: dict) -> str:
    parts = [_s(row.get(k)) for k in ("SURNAME", "FIRSTNAME", "PATRONYMIC")]
    return " ".join([p for p in parts if p]).strip()


# ─── Сборка структур ────────────────────────────────────────────────────────

def _build_education_entry(row: dict) -> dict | None:
    """Возвращает одну запись об образовании или None если пусто."""
    e = {
        "level":            _s(row.get("EDUCATION")),
        "institution":      _s(row.get("INSTITUTION")),
        "beg_date":         _s(row.get("EDUCATION_BEG_DATE")),
        "end_date":         _s(row.get("EDUCATION_END_DATE")),
        "diploma_num":      _s(row.get("DIPLOMA_NUM")),
        "specialty":        _s(row.get("SPECIALITY")),
        "degree":           _s(row.get("DEGREE")),
        "title":            _s(row.get("TITLE")),
        "specialnost_name": _s(row.get("SPECIALNOST")),
        "education_name_ru": _s(row.get("EDUCATION_NAME_RU")),
        "status":           _s(row.get("STATUS")),
        "vuz":              int(row["VUZ"]) if row.get("VUZ") is not None else None,
        "tipo":             int(row["TIPO"]) if row.get("TIPO") is not None else None,
        "srednee":          int(row["SREDNEE"]) if row.get("SREDNEE") is not None else None,
    }
    # если совсем пусто — None
    if not any(e[k] for k in ("level", "institution", "specialty", "education_name_ru")):
        return None
    return e


def _build_work_entry(row: dict) -> dict | None:
    e = {
        "country":   _s(row.get("LABOR_COUNTRY")),
        "city":      _s(row.get("CITY")),
        "company":   _s(row.get("COMPANY")),
        "post":      _s(row.get("POST")),
        "beg_date":  _s(row.get("LABOR_BEG_DATE")),
        "end_date":  _s(row.get("LABOR_END_DATE")),
        "experience": row.get("WORK_EXPERIENCE"),
    }
    if not any(e[k] for k in ("country", "city", "company", "post")):
        return None
    return e


def _build_military_entry(row: dict) -> dict | None:
    e = {
        "country":  _s(row.get("MILITARY_COUNTRY")),
        "branch":   _s(row.get("BRANCH")),
        "beg_date": _s(row.get("MILITARY_BEG_DATE")),
        "end_date": _s(row.get("MILITARY_END_DATE")),
    }
    if not any(e[k] for k in ("country", "branch")):
        return None
    return e


def _build_anketa(row: dict) -> dict:
    return {
        "trip_purpose":         _s(row.get("TRIP_PURPOSE")),
        "decision_num":         _s(row.get("DECISION_NUM")),
        "decision_date":        _s(row.get("DECISION_DATE")),
        "region":               _s(row.get("ANKETA_REGION")),
        "district":             _s(row.get("ANKETA_DISTRICT")),
        "address":              _s(row.get("ANKETA_ADDRESS")),
        "contact_person_iin":   _iin(row.get("CONTACT_PERSON_IIN")),
    }


def _build_kandas_status(row: dict) -> dict:
    return {
        "document_number":  _s(row.get("ORALMAN_DOCUMENT_NUMBER")),
        "status_date":      _s(row.get("DATE_ORALMAN_STATUS")),
        "region":           _s(row.get("KANDAS_REGION")),
        "district":         _s(row.get("KANDAS_DISTRICT")),
        "address":          _s(row.get("KANDAS_ADDRESS")),
        "quota_number":     _s(row.get("QUOTA_NUMBER")),
        "quota_date":       _s(row.get("QUOTA_DATE")),
    }


def _build_reg_address(row: dict) -> dict:
    return {
        "kato_reg":      _s(row.get("KATO_REG")),
        "kato_regname":  _s(row.get("KATO_REGNAME")),
        "kato_rai":      _s(row.get("KATO_RAI")),
        "kato_rainame":  _s(row.get("KATO_RAINAME")),
        "city":          _s(row.get("REG_ADDRESS_CITY")),
        "street":        _s(row.get("REG_ADDRESS_STREET")),
        "building":      _s(row.get("REG_ADDRESS_BUILDING")),
        "flat":          _s(row.get("REG_ADDRESS_FLAT")),
    }


def _build_labor_contract(row: dict) -> dict | None:
    e = {
        "org_bin":         _s(row.get("TD_ORG_BIN")),
        "org_name":        _s(row.get("TD_ORG_NAME")),
        "district":        _s(row.get("D_DISTRICT_NAME_RU")),
        "work_place_kato": _s(row.get("WORK_PLACE_D_KATO_NAME_RU")),
        "work_place":      _s(row.get("WORK_PLACE")),
        "part_time":       _s(row.get("D_PART_TIME_NAME_RU")),
        "remote_work":     _s(row.get("D_REMOTE_WORK_NAME_RU")),
        "working_hours":   _s(row.get("D_WORKING_HOURS_NAME_RU")),
        "position_code":   _s(row.get("D_POSITION_CODE")),
        "position_name":   _s(row.get("D_POSITION_NAME_RU")),
        "established_post": _s(row.get("ESTABLISHED_POST")),
        "contract_date":   _s(row.get("CONTRACT_DATE")),
        "termination_date": _s(row.get("TERMINATION_DATE")),
    }
    if not any(v for v in e.values()):
        return None
    return e


def _build_main_kandas(rows: list[dict], family: list[dict]) -> dict:
    """Собирает данные главного кандаса по нескольким строкам."""
    base = rows[0]

    # Списки: education / work / military
    edu_list = []
    work_list = []
    mil_list = []
    labor_contracts = []
    for r in rows:
        if e := _build_education_entry(r):
            if e not in edu_list:
                edu_list.append(e)
        if w := _build_work_entry(r):
            if w not in work_list:
                work_list.append(w)
        if m := _build_military_entry(r):
            if m not in mil_list:
                mil_list.append(m)
        if lc := _build_labor_contract(r):
            if lc not in labor_contracts:
                labor_contracts.append(lc)

    extra = {
        "nationality":     None,  # в новом xlsx нет — оставим если будет в будущем
        "family":          family,
        "anketa":          _build_anketa(base),
        "kandas_status":   _build_kandas_status(base),
        "reg_address":     _build_reg_address(base),
        "education":       edu_list,
        "work_history":    work_list,
        "military":        mil_list,
        "labor_contract":  labor_contracts[0] if labor_contracts else None,   # обычно 1 активный
        "occupation":      _s(base.get("OCCUPATIONA")),
        "family_type":     _s(base.get("FAMILY_TYPE")),
        "working":         int(base["WORKING"]) if base.get("WORKING") is not None else None,
        "avto":            int(base["AVTO"])    if base.get("AVTO")    is not None else None,
        "nedv":            int(base["NEDV"])    if base.get("NEDV")    is not None else None,
        "smz":             base.get("SMZ"),
        "sdu":             _s(base.get("SDU")),
        "cks_cat":         _s(base.get("SDU")),  # дублируем как cks_cat (legacy display)
        # Кол-во лиц, прикреплённых к поликлинике (новая колонка)
        "polyclinic":      base.get("прикрепление к поликлиннике"),
    }
    # Адрес проживания — используем KANDAS_REGION/DISTRICT + REG_ADDRESS
    oblast = _s(base.get("KANDAS_REGION"))
    raion  = _s(base.get("KANDAS_DISTRICT"))
    city   = _s(base.get("REG_ADDRESS_CITY")) or _s(base.get("KANDAS_ADDRESS"))
    street = _s(base.get("REG_ADDRESS_STREET"))
    house  = _s(base.get("REG_ADDRESS_BUILDING"))
    apt    = _s(base.get("REG_ADDRESS_FLAT"))

    dob = _s(base.get("DATE_BIRTH"))
    gender = _s(base.get("GENDER"))
    citizenship = _s(base.get("CITIZENSHIP_COUNTRY"))

    return {
        "fio":         _build_full_name(base),
        "iin":         _iin(base.get("IIN")),
        "dob":         dob,
        "age":         int(base["VOZRAST"]) if base.get("VOZRAST") is not None else None,
        "citizenship": citizenship,
        "gender":      gender,
        "oblast":      oblast,
        "raion":       raion,
        "city":        city,
        "street":      street,
        "house":       house,
        "apt":         apt,
        "phone":       None,   # в xlsx нет
        "extra":       extra,
    }


# ─── Main ───────────────────────────────────────────────────────────────────

async def main():
    xlsx_path = Path(sys.argv[1]) if len(sys.argv) > 1 else Path("/app/data/input/kandas_perso.xlsx")
    if not xlsx_path.exists():
        print(f"ERROR: file not found: {xlsx_path}")
        sys.exit(1)

    print(f"Reading {xlsx_path} ...")
    rows = _read_xlsx(xlsx_path)
    print(f"  parsed {len(rows)} rows")

    # Группируем по APP_ID
    by_app: dict = defaultdict(list)
    for r in rows:
        by_app[str(r["APP_ID"])].append(r)

    # Для каждой APP_ID-группы:
    #   - main kandas = ВСЕ строки где APPLICANT=1 (объединяем по IIN)
    #   - family members = строки где APPLICANT=0
    kandas_records: list[dict] = []
    for app_id, group in by_app.items():
        # Члены семьи (APPLICANT=0) — берём только role+fio+iin
        family_rows = [r for r in group if int(r.get("APPLICANT") or 0) == 0]
        family: list[dict] = []
        seen_family = set()
        for fr in family_rows:
            iin = _iin(fr.get("IIN"))
            if iin in seen_family:
                continue
            seen_family.add(iin)
            family.append({
                "role": _s(fr.get("RELATION_DEGREE")) or "Член семьи",
                "fio":  _build_full_name(fr),
                "iin":  iin,
            })

        # Главные (APPLICANT=1) — группируем по IIN, объединяем все строки
        main_rows = [r for r in group if int(r.get("APPLICANT") or 0) == 1]
        if not main_rows:
            print(f"  APP_ID={app_id}: нет главного кандаса (APPLICANT=1), пропуск")
            continue
        by_iin: dict = defaultdict(list)
        for mr in main_rows:
            iin = _iin(mr.get("IIN"))
            by_iin[iin].append(mr)
        for iin, rs in by_iin.items():
            kandas_records.append(_build_main_kandas(rs, family))

    print(f"\nГотово к загрузке: {len(kandas_records)} главных кандасов")

    # Upsert по IIN
    inserted = updated = 0
    async with AsyncSessionLocal() as db:
        for data in kandas_records:
            iin = data["iin"]
            existing = None
            if iin:
                res = await db.execute(select(Kandas).where(Kandas.iin == iin))
                existing = res.scalar_one_or_none()
            if existing:
                for k, v in data.items():
                    setattr(existing, k, v)
                updated += 1
                print(f"  Updated: {data['fio']} ({iin})")
            else:
                db.add(Kandas(**data))
                inserted += 1
                print(f"  Inserted: {data['fio']} ({iin})")
        await db.commit()

    print(f"\nDone: {inserted} inserted, {updated} updated.")


if __name__ == "__main__":
    asyncio.run(main())
