"""
Импортирует данные ЦОССУ из xlsx в таблицу cossu.

Запуск (один раз):
    docker compose exec backend bash -c "cd /app && python worker/seed_cossu.py /app/data/input/cossu.xlsx"

Если файл не указан — ищем по умолчанию /app/data/input/cossu.xlsx.

Логика:
- Каждая строка xlsx = одно отделение (одна запись в таблице cossu).
- branch_id уникален в рамках набора (внешний ID из колонки branch_ids).
- При повторном запуске: записи с тем же branch_id обновляются,
  записи без branch_id вставляются как новые (не дедуплицируются).
"""
import asyncio
import sys
from pathlib import Path

sys.path.insert(0, "/app")

import openpyxl
from sqlalchemy import select

from app.db.session import AsyncSessionLocal
from app.db.models import Cossu


# Маппинг номера колонки → имя поля в модели (из xlsx)
COLUMN_MAP = {
    1:  "branch_id",          # branch_ids
    2:  "region",
    3:  "kato_region",
    4:  "rayon",
    5:  "kato_rayon",
    6:  "rayon2",
    7:  "kato_rayon2",
    8:  "additional_address", # additional_address_ru
    9:  "org_bin",
    10: "sobst",
    11: "org_name",
    12: "fulladdress",
    13: "otd_name",
    14: "otd_typ",
    15: "otd_podtyp",
    16: "fakt_koika_mesto",
    17: "residents_count",
    18: "queue_count",
}

INT_FIELDS = {"fakt_koika_mesto", "residents_count", "queue_count"}


def _clean(value, field: str):
    """Нормализуем значение под тип поля."""
    if value is None:
        return None
    if field in INT_FIELDS:
        try:
            return int(value)
        except (ValueError, TypeError):
            return None
    s = str(value).strip()
    return s if s else None


def parse_xlsx(path: Path) -> list[dict]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    rows: list[dict] = []
    for r in range(2, ws.max_row + 1):
        data: dict = {}
        for col, field in COLUMN_MAP.items():
            data[field] = _clean(ws.cell(r, col).value, field)
        # Пропускаем строки без org_bin (битые)
        if not data.get("org_bin"):
            continue
        rows.append(data)
    return rows


async def main():
    xlsx_path = Path(sys.argv[1]) if len(sys.argv) > 1 else Path("/app/data/input/cossu.xlsx")
    if not xlsx_path.exists():
        print(f"ERROR: file not found: {xlsx_path}")
        sys.exit(1)

    print(f"Reading {xlsx_path} ...")
    rows = parse_xlsx(xlsx_path)
    print(f"  parsed {len(rows)} rows")

    inserted = updated = 0
    async with AsyncSessionLocal() as db:
        for data in rows:
            bid = data.get("branch_id")
            existing = None
            if bid:
                res = await db.execute(select(Cossu).where(Cossu.branch_id == bid))
                existing = res.scalar_one_or_none()

            if existing:
                for k, v in data.items():
                    setattr(existing, k, v)
                updated += 1
            else:
                db.add(Cossu(**data))
                inserted += 1
        await db.commit()

    print(f"\nDone: {inserted} inserted, {updated} updated.")


if __name__ == "__main__":
    asyncio.run(main())
