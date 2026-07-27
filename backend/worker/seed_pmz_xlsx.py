"""Импорт обладателей статуса постоянного резидента (ПМЖ) в таблицу kandas
с kind='pmz'. Отдельная сущность того же реестра — на фронте показывается
переключателем вместо кандасов.

Файл плоский: одна строка = один резидент (без семей/APPLICANT), шапка в
строке 1, данные со строки 2.

Запуск:
    docker compose cp pmz_20260727.xlsx backend:/app/data/input/pmz.xlsx
    docker compose exec backend python worker/seed_pmz_xlsx.py /app/data/input/pmz.xlsx
"""
from __future__ import annotations

import asyncio
import sys
from datetime import date, datetime
from pathlib import Path

sys.path.insert(0, "/app")

import openpyxl
from sqlalchemy import select

from app.db.session import AsyncSessionLocal
from app.db.models import Kandas


def _s(v) -> str | None:
    if v is None:
        return None
    s = str(v).strip()
    if not s or s in ("-", "None"):
        return None
    return s


def _iin(v) -> str | None:
    if v is None:
        return None
    s = str(v).strip()
    if "." in s:
        try:
            s = str(int(float(s)))
        except ValueError:
            pass
    s = s.split(".")[0]
    if s.isdigit() and len(s) == 11:
        s = "0" + s
    return s if s and s.isdigit() else (s or None)


def _int(v) -> int | None:
    if v is None or str(v).strip() in ("", "None", "-"):
        return None
    try:
        return int(float(v))
    except (ValueError, TypeError):
        return None


def _date(v) -> str | None:
    """Дату → 'дд.мм.гггг'. Принимает datetime/date/строку."""
    if v is None:
        return None
    if isinstance(v, (datetime, date)):
        return v.strftime("%d.%m.%Y")
    s = str(v).strip()
    return s or None


def _age(birth) -> int | None:
    if not isinstance(birth, (datetime, date)):
        return None
    t = date.today()
    return t.year - birth.year - ((t.month, t.day) < (birth.month, birth.day))


def _read_xlsx(path: Path) -> list[dict]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    # Шапка — первая строка с IIN среди первых 3
    hdr = None
    for r in range(1, 4):
        vals = {str(ws.cell(r, c).value).strip() for c in range(1, ws.max_column + 1) if ws.cell(r, c).value}
        if "IIN" in vals:
            hdr = r
            break
    if hdr is None:
        raise ValueError("не нашёл строку заголовков (IIN)")
    headers = {c: ws.cell(hdr, c).value for c in range(1, ws.max_column + 1) if ws.cell(hdr, c).value}
    rows = []
    for r in range(hdr + 1, ws.max_row + 1):
        d = {name: ws.cell(r, c).value for c, name in headers.items()}
        if _iin(d.get("IIN")):
            rows.append(d)
    return rows


def _build_pmz(row: dict) -> dict:
    fio = " ".join(p for p in (_s(row.get("SURNAME")), _s(row.get("FIRSTNAME")),
                               _s(row.get("SECONDNAME"))) if p)
    gid = _int(row.get("GENDER_ID"))
    gender = "М" if gid == 1 else "Ж" if gid == 2 else None

    # Образование одной записью
    vuz, tipo, srednee = _int(row.get("VUZ")), _int(row.get("TIPO")), _int(row.get("SREDNEE"))
    lvl = [n for n, f in (("Высшее", vuz), ("ТиПО", tipo), ("Среднее", srednee)) if f]
    edu = {
        "level":             ", ".join(lvl) or None,
        "education_name_ru": _s(row.get("EDUCATION_NAME_RU")),
        "specialnost_name":  _s(row.get("SPECIALNOST")),
        "status":            _s(row.get("STATUS")),
    }
    education = [edu] if any(edu.values()) else []

    # Трудовой договор (ЕСУТД)
    lc = {
        "org_bin":          _s(row.get("TD_ORG_BIN")),
        "org_name":         _s(row.get("TD_ORG_NAME")),
        "district":         _s(row.get("D_DISTRICT_NAME_RU")),
        "work_place_kato":  _s(row.get("WORK_PLACE_D_KATO_NAME_RU")),
        "work_place":       _s(row.get("WORK_PLACE")),
        "part_time":        _s(row.get("D_PART_TIME_NAME_RU")),
        "remote_work":      _s(row.get("D_REMOTE_WORK_NAME_RU")),
        "working_hours":    _s(row.get("D_WORKING_HOURS_NAME_RU")),
        "position_code":    _s(row.get("D_POSITION_CODE")),
        "position_name":    _s(row.get("D_POSITION_NAME_RU")),
        "established_post": _s(row.get("ESTABLISHED_POST")),
        "contract_date":    _date(row.get("CONTRACT_DATE")),
        "termination_date": _date(row.get("TERMINATION_DATE")),
    }
    labor_contract = lc if any(lc.values()) else None

    extra = {
        "nationality": _s(row.get("NATIONALITY")),
        "working":     _int(row.get("WORKING")),
        "avto":        _int(row.get("AVTO")),
        "nedv":        _int(row.get("NEDV")),
        "smz":         row.get("SMZ"),
        # Работодатель (текущий, из ORG_*)
        "bin_work":    _s(row.get("ORG_BIN")),
        "work_org":    _s(row.get("ORG_NAME")),
        "oked":        _s(row.get("VNAME_OKED")),
        # Статус резидента — даты
        "pmz_status": {
            "reg_date":     _date(row.get("REG_DATE")),
            "scoring_date": _date(row.get("SCORING_DATE")),
            "pmz_date":     _date(row.get("PMZ_DATE")),
        },
        "reg_address": {
            "kato_reg":     _s(row.get("KATO_REG")),
            "kato_regname": _s(row.get("KATO_REGNAME")),
            "kato_rai":     _s(row.get("KATO_RAI")),
            "kato_rainame": _s(row.get("KATO_RAINAME")),
            "city":         _s(row.get("REG_ADDRESS_CITY")),
            "street":       _s(row.get("REG_ADDRESS_STREET")),
            "building":     _s(row.get("REG_ADDRESS_BUILDING")),
            "flat":         _s(row.get("REG_ADDRESS_FLAT")),
        },
        "education":      education,
        "labor_contract": labor_contract,
    }

    return {
        "kind":        "pmz",
        "fio":         fio,
        "iin":         _iin(row.get("IIN")),
        "dob":         _date(row.get("BIRTHDATE")),
        "age":         _age(row.get("BIRTHDATE")),
        "citizenship": _s(row.get("CITIZANSHIP")),
        "gender":      gender,
        "oblast":      _s(row.get("REGION")) or _s(row.get("KATO_REGNAME")),
        "raion":       _s(row.get("KATO_RAINAME")),
        "city":        _s(row.get("REG_ADDRESS_CITY")) or _s(row.get("KATO_REGNAME")),
        "street":      _s(row.get("REG_ADDRESS_STREET")),
        "house":       _s(row.get("REG_ADDRESS_BUILDING")),
        "apt":         _s(row.get("REG_ADDRESS_FLAT")),
        "phone":       None,
        "extra":       extra,
    }


async def main():
    xlsx_path = Path(sys.argv[1]) if len(sys.argv) > 1 else Path("/app/data/input/pmz.xlsx")
    if not xlsx_path.exists():
        print(f"ERROR: file not found: {xlsx_path}")
        sys.exit(1)

    print(f"Reading {xlsx_path} ...")
    rows = _read_xlsx(xlsx_path)
    print(f"  parsed {len(rows)} резидентов")

    records = [_build_pmz(r) for r in rows if _build_pmz(r)["iin"]]
    print(f"Готово к загрузке: {len(records)} резидентов")

    inserted = updated = 0
    async with AsyncSessionLocal() as db:
        for data in records:
            iin = data["iin"]
            existing = None
            if iin:
                res = await db.execute(
                    select(Kandas).where(Kandas.iin == iin, Kandas.kind == "pmz")
                )
                existing = res.scalar_one_or_none()
            if existing:
                for k, v in data.items():
                    setattr(existing, k, v)
                updated += 1
                print(f"  Updated: {data['fio']} ({iin})")
            else:
                db.add(Kandas(**data))
                inserted += 1
                print(f"  Inserted: {data['fio']} ({iin})")
        await db.commit()

    print(f"\nDone: {inserted} inserted, {updated} updated.")


if __name__ == "__main__":
    asyncio.run(main())
